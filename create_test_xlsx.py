"""
테스트용 XLSX 파일 생성 스크립트
셀 단위 저장 기능을 테스트하기 위한 간단한 엑셀 파일을 생성합니다.
"""

import openpyxl
from pathlib import Path

def create_test_xlsx():
    """테스트용 XLSX 파일 생성"""
    workbook = openpyxl.Workbook()
    
    # 기본 시트 이름 변경
    sheet = workbook.active
    sheet.title = "테스트데이터"
    
    # 테스트 데이터 입력
    test_data = [
        ["이름", "나이", "직업", "거주지"],
        ["김철수", 30, "개발자", "서울"],
        ["이영희", 25, "디자이너", "부산"],
        ["박민수", 35, "기획자", "대구"],
        ["정수연", 28, "마케터", "인천"],
        ["", "", "", ""],  # 빈 행
        ["제품명", "가격", "재고", "카테고리"],
        ["노트북", 1500000, 10, "전자제품"],
        ["마우스", 50000, 25, "전자제품"],
        ["키보드", 120000, 15, "전자제품"],
        ["모니터", 300000, 8, "전자제품"],
        ["", "", "", ""],  # 빈 행
        ["한국어 텍스트", "영어 텍스트", "숫자", "날짜"],
        ["안녕하세요", "Hello", 12345, "2024-01-01"],
        ["반갑습니다", "Nice to meet you", 67890, "2024-02-15"],
        ["감사합니다", "Thank you", 11111, "2024-03-20"],
    ]
    
    # 데이터 입력
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # 추가 시트 생성
    sheet2 = workbook.create_sheet("부서정보")
    dept_data = [
        ["부서명", "팀장", "인원수"],
        ["개발팀", "김개발", 12],
        ["디자인팀", "이디자인", 8],
        ["기획팀", "박기획", 6],
        ["마케팅팀", "정마케팅", 10],
    ]
    
    for row_idx, row_data in enumerate(dept_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            sheet2.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # 파일 저장
    file_path = Path("test_cells.xlsx")
    workbook.save(file_path)
    workbook.close()
    
    print(f"테스트 XLSX 파일 생성 완료: {file_path.absolute()}")
    print(f"시트 수: 2개 (테스트데이터, 부서정보)")
    print(f"총 셀 수: 약 40개 (빈 셀 제외)")
    
    return file_path

if __name__ == "__main__":
    create_test_xlsx()
